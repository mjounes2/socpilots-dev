"""
Multi-format file text extraction and chunking for SOCPilots Evidence Ingestion.
Supports: PDF, Excel/XLSX, CSV, TXT/LOG, JPG/PNG (OCR).
"""

import io
import logging
import re

log = logging.getLogger(__name__)

ALLOWED_EXTENSIONS = {'.pdf', '.xlsx', '.xls', '.csv', '.txt', '.log', '.jpg', '.jpeg', '.png'}
MAX_FILE_SIZE = 20 * 1024 * 1024  # 20 MB

# Magic-byte fingerprints for type validation
_MAGIC = {
    b'%PDF':               'application/pdf',
    b'PK\x03\x04':        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    b'\xff\xd8\xff':       'image/jpeg',
    b'\x89PNG\r\n\x1a\n': 'image/png',
}

ALLOWED_MIMES = {
    'application/pdf',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'application/vnd.ms-excel',
    'text/csv',
    'text/plain',
    'text/x-log',
    'image/jpeg',
    'image/png',
}


def _magic_mime(data: bytes) -> str | None:
    for sig, mime in _MAGIC.items():
        if data[:len(sig)] == sig:
            return mime
    try:
        data[:1024].decode('utf-8')
        return 'text/plain'
    except Exception:
        return None


def validate_file(data: bytes, filename: str) -> tuple[bool, str, str]:
    """Return (ok, detected_mime, error). Checks size, extension, magic bytes."""
    if len(data) > MAX_FILE_SIZE:
        mb = len(data) // (1024 * 1024)
        return False, '', f'File too large: {mb}MB (max 20MB)'

    ext = ('.' + filename.rsplit('.', 1)[-1].lower()) if '.' in filename else ''
    if ext not in ALLOWED_EXTENSIONS:
        return False, '', f'Unsupported file extension: {ext}'

    mime = _magic_mime(data)
    if mime is None or mime not in ALLOWED_MIMES:
        return False, '', f'File type rejected by magic-byte check (detected: {mime})'

    return True, mime, ''


def extract_text(data: bytes, mime: str, filename: str) -> str:
    """Dispatch to the correct extractor; return empty string on failure."""
    try:
        if mime == 'application/pdf':
            return _pdf(data)
        if mime in ('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    'application/vnd.ms-excel'):
            return _excel(data)
        if mime == 'text/csv':
            return _csv(data)
        if mime in ('text/plain', 'text/x-log'):
            return data.decode('utf-8', errors='replace')
        if mime in ('image/jpeg', 'image/png'):
            return _ocr(data)
    except Exception as exc:
        log.warning('Text extraction failed for %s (%s): %s', filename, mime, exc)
    return ''


def _pdf(data: bytes) -> str:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(data))
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ''
        if text.strip():
            pages.append(f'[Page {i + 1}]\n{text}')
    return '\n\n'.join(pages)


def _excel(data: bytes) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 1000:
                rows.append('[… truncated at 1000 rows]')
                break
            cells = ['' if c is None else str(c) for c in row]
            if any(c.strip() for c in cells):
                rows.append('\t'.join(cells))
        if rows:
            sheets.append(f'[Sheet: {name}]\n' + '\n'.join(rows))
    return '\n\n'.join(sheets)


def _csv(data: bytes) -> str:
    import csv
    text = data.decode('utf-8', errors='replace')
    rows = []
    for i, row in enumerate(csv.reader(io.StringIO(text))):
        if i >= 2000:
            rows.append('[… truncated at 2000 rows]')
            break
        rows.append('\t'.join(row))
    return '\n'.join(rows)


def _ocr(data: bytes) -> str:
    try:
        import pytesseract
        from PIL import Image
        img = Image.open(io.BytesIO(data))
        return pytesseract.image_to_string(img)
    except ImportError:
        log.warning('pytesseract/Pillow not installed — OCR unavailable')
        return ''


def chunk_text(text: str, chunk_size: int = 600, overlap: int = 60) -> list[str]:
    """Split text into overlapping character-level chunks for embedding."""
    text = re.sub(r'\s+', ' ', text).strip()
    if not text:
        return []
    chunks = []
    start = 0
    while start < len(text):
        chunk = text[start:start + chunk_size].strip()
        if chunk:
            chunks.append(chunk)
        start += chunk_size - overlap
    return chunks
